import os
from enum import Enum
from openpyxl import Workbook, load_workbook
import openpyxl.utils.cell
from openpyxl.cell import read_only


def handleups(fullpath, select_filename, tipps_value, root, progressbar):
    # openpyxl 索引从1开始
    col_list = []
    col_list1 = [('发票号码', 6, 1),
                 ('发票金额', 11, 2),
                 ('发票日期', 5, 3),
                 ('客户', 23, 4),
                 ('货件参考编码 1', 16, 5),
                 ('主单号', 14, 6),
                 ('子单号', 21, 7),
                 ('包裹数', 19, 8),
                 ('计费重量KG', 29, 9),
                 ('收件人国家', 82, 10),
                 ('K5打单金额', 25, 11)]
    col_list2 = []  # 费用类型
    col_list3_part = [('交易货币代码', 51),
                      ('主单号总金额', 56),
                      ('发件人公司名称', 68),
                      ('发件人邮编', 73),
                      ('发件人国家或地区', 74),
                      ('收件人公司名称', 76),
                      ('收件人邮编', 81),
                      ('收件人所在国家或地区', 82)]
    col_list3 = []
    col_fee_type = 46  # index从1开始
    col_cost = 53  # 附加费

    print('handlesups path:', fullpath)

    # df = pd.read_excel(str(fullpath), sheet_name=0)  # 打开一个xlsx文件
    # nrows = df.shape[0]  # 返回df的行数
    # ncols = df.shape[1]  # 列数

    print(os.path.splitext(select_filename)[1])
    if os.path.splitext(select_filename)[1] == '.xlsm':
        wb = load_workbook(fullpath, read_only=False, keep_vba=True)  # 坑！！！xlsm文件True，xlsx文件False
    else:
        wb = load_workbook(fullpath, read_only=False, keep_vba=False)

    tipps_value.set('提示:文件正在处理...')
    root.update()

    ws = wb[wb.sheetnames[0]]
    maxrow = ws.max_row

    # 获取费用类型数据
    tuple_fee_type = ws['AT']
    list_fee_type = []
    for i in range(len(tuple_fee_type)):
        list_fee_type.append(list(tuple_fee_type)[i].value)

    # 去重
    list_fee_type = list(set(list_fee_type))

    # 将费用类型赋值给col_list2
    m = 1
    for i in range(len(list_fee_type)):
        if list_fee_type[i] == '费用描述':
            continue
        temp_list = [list_fee_type[i], col_fee_type, len(col_list1) + m]
        col_list2.append(tuple(temp_list))
        m = m + 1
    # 添加col_list3在新表中的列索引
    for i in range(len(col_list3_part)):
        temp_list = list(col_list3_part[i])
        temp_list.append(len(col_list1) + len(col_list2) + i + 1)
        col_list3.append(tuple(temp_list))

    print(col_list2)

    # 合并col_list
    col_list.extend(col_list1)
    col_list.extend(col_list2)
    col_list.extend(col_list3)

    # 新建sheet
    try:
        wb.get_sheet_by_name('已处理')
        wb.remove(wb['已处理'])
        print('删除之前生成的sheet')
    except KeyError:
        pass

    print('新建sheet')
    wb.create_sheet('已处理')
    new_ws = wb.get_sheet_by_name('已处理')
    for i in range(len(col_list)):
        new_ws.cell(1, col_list[i][2]).value = col_list[i][0]
    # 按行读取，第一行是标题行
    for i in range(2, maxrow+1):

        #  更新界面processbar
        progressbar['value'] = (i/maxrow)*100
        root.update()

        new_max_row = new_ws.max_row
        if ws.cell(i, col_list[5][1]).value == ws.cell(i - 1, col_list[5][1]).value:
            # 如果主单号相同
            if ws.cell(i, col_list[6][1]).value != ws.cell(i - 1, col_list[6][1]).value:
                # 如果子单号不同
                new_ws.cell(new_max_row, col_list[6][2]).value = new_ws.cell(new_max_row, col_list[6][2]).value + ',' + ws.cell(i, col_list[6][1]).value
            for fee in col_list2:
                #  附加费用遍历
                if ws.cell(i, fee[1]).value == fee[0]:
                    if new_ws.cell(new_max_row, fee[2]).value is None:
                        new_ws.cell(new_max_row, fee[2]).value = 0
                    if ws.cell(i, col_cost).value is None:
                        ws.cell(i, col_cost).value = 0
                    # print(new_ws.cell(new_max_row, fee[2]).value)
                    # print(ws.cell(i, col_cost).value)
                    try:
                        new_ws.cell(new_max_row, fee[2]).value = float(new_ws.cell(new_max_row, fee[2]).value) + float(ws.cell(i, col_cost).value)
                    except ValueError:
                        print(new_max_row)
                        print(fee[2])
                        print(new_ws.cell(new_max_row, fee[2]).value)
                        print(ws.cell(i, col_cost).value)
            if ws.cell(i, col_list3[1][1]).value != 0:
                # 主单号总金额
                new_ws.cell(new_max_row, col_list3[1][2]).value = float(new_ws.cell(new_max_row, col_list3[1][2]).value) + float(ws.cell(i, col_list3[1][1]).value)
                # print(new_ws.cell(new_max_row, col_list1[5][2]).value, " is ", float(ws.cell(i, col_list3[1][1]).value))
        else:
            for x in range(len(col_list1)):
                new_ws.cell(new_max_row + 1, col_list1[x][2]).value = ws.cell(i, col_list1[x][1]).value
            for fee in col_list2:
                if ws.cell(i, fee[1]).value == fee[0]:
                    new_ws.cell(new_max_row + 1, fee[2]).value = ws.cell(i, col_cost).value
            for x in range(len(col_list3)):
                new_ws.cell(new_max_row + 1, col_list3[x][2]).value = ws.cell(i, col_list3[x][1]).value
            # print(ws.cell(i, 53).value)
    try:
        wb.save(fullpath)
        wb.close()
    except PermissionError:
        tipps_value.set('文件已被占用,请先关闭')
        root.update()
        return 'please close file'

    return 'finish'
