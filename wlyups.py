import os
from openpyxl import load_workbook



def handle_wlyups(fullpath, select_filename, tipps_value, root, progressbar):
    # openpyxl 索引从1开始
    col_list = []
    col_list1 = [('发票号码', 6, 1),
                 ('发票金额', 11, 2),
                 ('发票日期', 5, 3),
                 ('客户', 23, 4),
                 ('收件人国家', 24, 5),
                 ('打单金额', 25, 6),
                 ('货件参考编码 1', 16, 7),
                 ('主单号', 14, 8),
                 ('子单号', 21, 9),
                 ('包裹数', 19, 10),
                 ('计费重量KG', 29, 11)]

    col_list2 = []  # 所有费用类型
    col_list3_part = [('交易货币代码', 51),
                      ('主单号总金额', 56),
                      ('发件人公司名称', 68),
                      ('发件人邮编', 73),
                      ('发件人国家或地区', 74),
                      ('收件人公司名称', 76),
                      ('收件人邮编', 81),
                      ('收件人所在国家或地区', 82)]
    col_list3 = []
    col_fee_type = 46  # 费用描述所在旧表中的列索引.index从1开始
    col_cost = 53  # 附加费

    # print('handlesups path:', fullpath)

    # df = pd.read_excel(str(fullpath), sheet_name=0)  # 打开一个xlsx文件
    # nrows = df.shape[0]  # 返回df的行数
    # ncols = df.shape[1]  # 列数

    # print(os.path.splitext(select_filename)[1])
    if os.path.splitext(select_filename)[1] == '.xlsm':
        wb = load_workbook(fullpath, read_only=False, keep_vba=True)  # 坑！！！xlsm文件True，xlsx文件False
    else:
        wb = load_workbook(fullpath, read_only=False, keep_vba=False)

    tipps_value.set('提示:文件正在处理...')
    root.update()

    ws = wb[wb.sheetnames[0]]
    maxrow = ws.max_row

    # 获取费用类型数据
    tuple_fee_type = ws['AT']  # excel AT 列
    list_fee_type = []
    for i in range(len(tuple_fee_type)):
        list_fee_type.append(list(tuple_fee_type)[i].value)

    # 去重
    list_fee_type = list(set(list_fee_type))

    # 将费用类型赋值给col_list2，所有费用名称列表
    m = 1
    for i in range(len(list_fee_type)):
        if list_fee_type[i] == '费用描述':
            continue
        temp_list = [list_fee_type[i], col_fee_type, len(col_list1) + m]  # [(费用描述),(费用描述所在旧表中的列索引),(费用描述所在新表中的列索引)]
        col_list2.append(tuple(temp_list))
        m = m + 1

    # 添加col_list3在新表中的列索引。 col_list3 = [col_list3_part,新表索引]，即转为col_list1的样式
    for i in range(len(col_list3_part)):
        temp_list = list(col_list3_part[i])
        temp_list.append(len(col_list1) + len(col_list2) + i + 1)
        col_list3.append(tuple(temp_list))

    # print(col_list2)

    # 合并col_list
    col_list.extend(col_list1)
    col_list.extend(col_list2)
    col_list.extend(col_list3)

    # 新建sheet
    try:
        wb.get_sheet_by_name('已处理')
        wb.remove(wb['已处理'])
        # print('删除之前生成的sheet')
    except KeyError:
        pass

    # print('新建sheet')
    wb.create_sheet('已处理')
    new_ws = wb.get_sheet_by_name('已处理')

    # 判断是否是对应的文件
    if ws['U1'].value != '追踪编号':
        tipps_value.set('文件格式不正确')
        root.update()
        return 'file false'

    # 新表第一行，添加列名
    for i in range(len(col_list)):
        new_ws.cell(1, col_list[i][2]).value = col_list[i][0]

    # 按行读取旧表，第一行是标题行，从第二行开始。依次处理数据后，添加到新表。
    for i in range(2, maxrow+1):

        #  更新界面processbar
        progressbar['value'] = (i/maxrow)*100
        root.update()

        new_max_row = new_ws.max_row  # 新表最大行

        main_order_no_old_col = col_list[7][1]  # 主单号在旧表的列索引
        sub_order_no_old_col = col_list[8][1]  # 子单号在旧表的列索引
        sub_order_no_new_col = col_list[8][2]  # 子单号在新表的列索引
        main_order_cost_old_col = col_list3[1][1]  # 主单号总金额在旧表的列索引
        main_order_cost_new_col = col_list3[1][2]  # 主单号总金额在新表的列索引

        # 如果主单号与前一行相同
        if ws.cell(i, main_order_no_old_col).value == ws.cell(i - 1, main_order_no_old_col).value:

            # 如果子单号不同，拼接子单号
            if ws.cell(i, sub_order_no_old_col).value != ws.cell(i - 1, sub_order_no_old_col).value:

                new_ws.cell(new_max_row, sub_order_no_new_col).value = new_ws.cell(new_max_row, sub_order_no_new_col).value + ',' + ws.cell(i, sub_order_no_old_col).value

            for fee_1 in col_list2:
                #  附加费用遍历
                # print(fee)

                # 如果匹配到费用名称
                if ws.cell(i, fee_1[1]).value == fee_1[0]:
                    # print("主单号与上一行主单号相同")

                    if new_ws.cell(new_max_row, fee_1[2]).value is None:
                        new_ws.cell(new_max_row, fee_1[2]).value = 0
                    if ws.cell(i, col_cost).value is None:
                        ws.cell(i, col_cost).value = 0

                    try:
                        new_ws.cell(new_max_row, fee_1[2]).value = float(new_ws.cell(new_max_row, fee_1[2]).value) + float(ws.cell(i, col_cost).value)
                        # print(fee_1[0])
                        # print(new_ws.cell(new_max_row, fee_1[2]).value)

                    except ValueError:
                        pass
                        # print(new_max_row)
                        # print(fee[2])
                        # print(new_ws.cell(new_max_row, fee[2]).value)
                        # print(ws.cell(i, col_cost).value)

        # 如果旧表中主单号与上一行主单号不同，则新表中新加一行
        else:
            # print("主单号与上一行主单号不同")

            for x in range(len(col_list1)):
                new_ws.cell(new_max_row + 1, col_list1[x][2]).value = ws.cell(i, col_list1[x][1]).value

            for fee_2 in col_list2:
                #  如果匹配到收费名称
                if ws.cell(i, fee_2[1]).value == fee_2[0]:
                    if ws.cell(i, col_cost).value is None:
                        new_ws.cell(new_max_row + 1, fee_2[2]).value = 0
                    else:
                        new_ws.cell(new_max_row + 1, fee_2[2]).value = ws.cell(i, col_cost).value
                    # print(new_ws.cell(1, fee_2[2]).value)
                    # print(new_ws.cell(new_max_row + 1, fee_2[2]).value)

                #  没匹配到，先将费用置为0
                else:
                    new_ws.cell(new_max_row + 1, fee_2[2]).value = 0

            for x in range(len(col_list3)):
                new_ws.cell(new_max_row + 1, col_list3[x][2]).value = ws.cell(i, col_list3[x][1]).value

        # 主单号总金额
        new_max_row = new_ws.max_row  # 新表最大行
        if new_max_row > 1:
            # print("主单号总金额")

            main_cost = 0
            for main_cost_i in range(len(col_list1)+1, len(col_list1)+len(col_list2)+1):
                # print(new_ws.cell(1, main_cost_i).value)
                # print(new_ws.cell(new_max_row, main_cost_i).value)
                if "Tax" not in new_ws.cell(1, main_cost_i).value:
                    main_cost = float(new_ws.cell(new_max_row, main_cost_i).value)+main_cost
            new_ws.cell(new_max_row, main_order_cost_new_col).value = main_cost

    try:
        wb.save(fullpath)
        wb.close()
    except PermissionError:
        tipps_value.set('文件已被占用,请先关闭')
        root.update()
        return 'please close file'

    return 'finish'
