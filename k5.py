import os
from openpyxl import load_workbook


def copy_row(row_no, new_row_no, ws, new_ws):
    for cell in ws[row_no]:
        new_ws.cell(row=new_row_no, column=cell.column, value=cell.value)


def handle_k5(fullpath, select_filename, tipps_value, root, progressbar):
    if os.path.splitext(select_filename)[1] == '.xlsm':
        wb = load_workbook(fullpath, read_only=False, keep_vba=True)  # 坑！！！xlsm文件True，xlsx文件False
    else:
        wb = load_workbook(fullpath, read_only=False, keep_vba=False)

    tipps_value.set('提示:文件正在处理...')
    root.update()

    ws = wb[wb.sheetnames[0]]
    maxrow = ws.max_row

    # 新建sheet
    try:
        wb.get_sheet_by_name('已处理')
        wb.remove(wb['已处理'])
        # print('删除之前生成的sheet')
    except KeyError:
        pass

    # print('新建sheet')
    wb.create_sheet('已处理')
    new_ws = wb.get_sheet_by_name('已处理')

    # 判断是否是对应的文件
    if ws['F1'].value != '子单号':
        tipps_value.set('文件格式不正确')
        root.update()
        return 'file false'

    # 进行数据处理
    # copy_row(1, 1, ws, new_ws)
    new_ws.cell(row=1, column=1, value=ws[1][0].value)
    new_ws.cell(row=1, column=2, value=ws[1][1].value)
    new_ws.cell(row=1, column=3, value=ws[1][2].value)
    new_ws.cell(row=1, column=4, value=ws[1][3].value)
    new_ws.cell(row=1, column=5, value=ws[1][4].value)
    new_ws.cell(row=1, column=6, value=ws[1][5].value)
    new_ws.cell(row=1, column=7, value=ws[1][8].value)
    new_ws.cell(row=1, column=8, value=ws[1][11].value)
    new_ws.cell(row=1, column=9, value=ws[1][46].value)
    new_ws.cell(row=1, column=10, value=ws[1][47].value)
    new_ws.cell(row=1, column=11, value=ws[1][49].value)
    new_ws.cell(row=1, column=12, value=ws[1][50].value)
    new_ws.cell(row=1, column=13, value=ws[1][51].value)

    for i in range(2, maxrow+1):

        #  更新界面processbar
        progressbar['value'] = (i / maxrow) * 100
        root.update()

        old_row = ws[i]
        tracking_nos = old_row[5].value
        tracking_nos_list = list(tracking_nos.split(','))
        # print(tracking_nos_list)
        for tracking_no in tracking_nos_list:
            new_maxrow = new_ws.max_row
            # copy_row(i, new_maxrow + 1, ws, new_ws)
            new_ws.cell(row=new_maxrow + 1, column=1, value=old_row[0].value)
            new_ws.cell(row=new_maxrow + 1, column=2, value=old_row[1].value)
            new_ws.cell(row=new_maxrow + 1, column=3, value=old_row[2].value)
            new_ws.cell(row=new_maxrow + 1, column=4, value=old_row[3].value)
            new_ws.cell(row=new_maxrow + 1, column=5, value=old_row[4].value)
            new_ws.cell(row=new_maxrow + 1, column=6, value=tracking_no)
            new_ws.cell(row=new_maxrow + 1, column=7, value=old_row[8].value)
            new_ws.cell(row=new_maxrow + 1, column=8, value=old_row[11].value)
            new_ws.cell(row=new_maxrow + 1, column=9, value=old_row[46].value)
            new_ws.cell(row=new_maxrow + 1, column=10, value=old_row[47].value)
            new_ws.cell(row=new_maxrow + 1, column=11, value=old_row[49].value)
            new_ws.cell(row=new_maxrow + 1, column=12, value=old_row[50].value)
            new_ws.cell(row=new_maxrow + 1, column=13, value=old_row[51].value)
    try:
        wb.save(fullpath)
        wb.close()
    except PermissionError:
        tipps_value.set('文件已被占用,请先关闭')
        root.update()
        return 'please close file'

    return 'finish'
