import os

import openpyxl
from openpyxl import load_workbook


def load_workbook_price(filename, sheet_name):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    return sheet


# 提取数据并返回列表的列表形式
def extract_data(sheet):
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data[1:], data[0]  # 假设第一行是表头，所以从第二行开始提取数据


# 模拟查询函数：
def query_data(data, business_type, customer_code, weight_country):
    rows, headers = data
    for row in rows:
        if row[0] == business_type and row[1] == customer_code:
            if isinstance(weight_country, float):
                weight_range = row[2].split('-')
                if float(weight_range[0]) < weight_country <= float(weight_range[1]):
                    return row[3], row[4]
            else:
                if row[2] == weight_country:
                    return row[3], row[4]
    return "查询不到报价"


def handle_dhl(fullpath, select_filename, tipps_value, root, progressbar, dhl_price_fullpath, dhl_price):
    # 获取dhl价格---开始
    # 加载工作簿和工作表

    wb_price = load_workbook(dhl_price_fullpath)
    sheet_price = wb_price[wb_price.sheetnames[0]]
    data_price = extract_data(sheet_price)
    # 获取dhl价格---结束

    # openpyxl 索引从1开始
    col_list = []
    col_list1 = [('Invoice Nr', 4, 1),
                 ('Invoice Date', 3, 2),
                 ('Pu Date', 12, 3),
                 ('Identcode', 13, 4),
                 ('货主代码', 14, 5),
                 ('发货渠道', 15, 6),
                 ('目的地国家', 16, 7),
                 ('K5打单金额', 17, 8),
                 ('件数', 18, 9),
                 ('Shippers Reference', 19, 10),
                 ('Pcs', 20, 11),
                 ('Wgt', 22, 12),
                 ('Wgt_Abr', 23, 13)]
    col_list2 = []  # 费用类型
    col_list3_part = [('Total', 27)]
    col_list3 = []
    col_fee_type = 11  # 原表费用类型位置，index从1开始
    col_cost = 27  # 总费用
    col_old_charge_amount = 24  # 基础费
    col_old_extra_charge_amount = 25  # 附加费

    if os.path.splitext(select_filename)[1] == '.xlsm':
        wb = load_workbook(fullpath, read_only=False, keep_vba=True)  # 坑！！！xlsm文件True，xlsx文件False
    else:
        wb = load_workbook(fullpath, read_only=False, keep_vba=False)

    tipps_value.set('提示:文件正在处理...')
    root.update()

    ws = wb[wb.sheetnames[0]]
    maxrow = ws.max_row

    # 获取费用类型数据
    tuple_fee_type = ws['K']
    list_fee_type = []
    for i in range(len(tuple_fee_type)):
        if len(str(list(tuple_fee_type)[i].value)) < 9:
            list_fee_type.append(list(tuple_fee_type)[i].value)

    # 去重
    list_fee_type = list(set(list_fee_type))

    list_fee_type.remove("Prod")
    list_fee_type.append("Charge Amount")
    list_fee_type.append("Extra Charge Amount")

    # list_fee_type = ["Charge Amount", "Extra Charge Amount"]

    # 将费用类型赋值给col_list2
    m = 1
    for i in range(len(list_fee_type)):
        if list_fee_type[i] == '费用描述':
            continue
        temp_list = [list_fee_type[i], col_fee_type, len(col_list1) + m]
        col_list2.append(tuple(temp_list))
        m = m + 1
    # 添加col_list3在新表中的列索引
    for i in range(len(col_list3_part)):
        temp_list = list(col_list3_part[i])
        temp_list.append(len(col_list1) + len(col_list2) + i + 1)
        col_list3.append(tuple(temp_list))

    # print(col_list2)
    # print(col_list3)
    # 合并col_list
    col_list.extend(col_list1)
    col_list.extend(col_list2)
    col_list.extend(col_list3)

    # 新建sheet
    try:
        wb.get_sheet_by_name('已处理')
        wb.remove(wb['已处理'])
        # print('删除之前生成的sheet')
    except KeyError:
        pass

    # print('新建sheet')
    wb.create_sheet('已处理')
    new_ws = wb.get_sheet_by_name('已处理')

    # 判断是否是对应的文件
    # print(ws['M1'].value)
    if ws['M1'].value != 'Identcode':
        tipps_value.set('文件格式不正确')
        root.update()
        return 'file false'

    for i in range(len(col_list)):
        new_ws.cell(1, col_list[i][2]).value = col_list[i][0]
    # 按行读取，第一行是标题行
    for i in range(2, maxrow + 1):
        # print("i="+str(i))
        #  更新界面processbar
        progressbar['value'] = (i / maxrow) * 100
        root.update()

        new_max_row = new_ws.max_row

        if ws.cell(i, col_list[3][1]).value == ws.cell(i - 1, col_list[3][1]).value:
            # 如果col_list[3] = Identcode 单号相同
            if ws.cell(i, col_list[3][1]).value is not None:

                for fee in col_list2:
                    #  附加费用col_list2遍历

                    if fee[0] == "Charge Amount":
                        if ws.cell(i, col_old_charge_amount).value is not None:

                            if new_ws.cell(new_max_row, fee[2]).value is None:
                                new_ws.cell(new_max_row, fee[2]).value = 0
                            if ws.cell(i, col_cost).value is None:
                                ws.cell(i, col_cost).value = 0

                            try:
                                new_ws.cell(new_max_row, fee[2]).value = float(
                                    new_ws.cell(new_max_row, fee[2]).value) + float(ws.cell(i, col_cost).value)
                                print(new_ws.cell(new_max_row, fee[2]).value)
                            except ValueError:
                                pass
                                # print(new_max_row)
                                # print(fee[2])
                                # print(new_ws.cell(new_max_row, fee[2]).value)
                                # print(ws.cell(i, col_cost).value)
                    if fee[0] == "Extra Charge Amount":
                        if ws.cell(i, col_old_extra_charge_amount).value is not None:
                            if new_ws.cell(new_max_row, fee[2]).value is None:
                                new_ws.cell(new_max_row, fee[2]).value = 0
                            if ws.cell(i, col_cost).value is None:
                                ws.cell(i, col_cost).value = 0

                            try:
                                new_ws.cell(new_max_row, fee[2]).value = float(
                                    new_ws.cell(new_max_row, fee[2]).value) + float(ws.cell(i, col_cost).value)
                            except ValueError:
                                pass
                                # print(new_max_row)
                                # print(fee[2])
                                # print(new_ws.cell(new_max_row, fee[2]).value)
                                # print(ws.cell(i, col_cost).value)
                    # print("total:"+str(col_list3[0][1]))

                    # 附加费用

                    if fee[0] == ws.cell(i, 11).value:
                        new_ws.cell(new_max_row, fee[2]).value = ws.cell(i, col_cost).value

                if ws.cell(i, col_list3[0][1]).value != 0:
                    # 主单号总金额

                    new_ws.cell(new_max_row, col_list3[0][2]).value = float(
                        new_ws.cell(new_max_row, col_list3[0][2]).value) + float(ws.cell(i, col_list3[0][1]).value)
                    # print(new_ws.cell(new_max_row, col_list3[0][2]).value, " + ", float(ws.cell(i, col_list3[0][1]).value))


            else:
                for x in range(len(col_list1)):
                    new_ws.cell(new_max_row + 1, col_list1[x][2]).value = ws.cell(i, col_list1[x][1]).value
                for fee in col_list2:
                    if fee[0] == "Charge Amount":
                        if ws.cell(i, col_old_charge_amount).value is not None:
                            if ws.cell(i, col_old_charge_amount).value != '':
                                new_ws.cell(new_max_row + 1, fee[2]).value = ws.cell(i, col_cost).value
                    if fee[0] == "Extra Charge Amount":
                        if ws.cell(i, col_old_extra_charge_amount).value is not None:
                            if ws.cell(i, col_old_extra_charge_amount).value != '':
                                new_ws.cell(new_max_row + 1, fee[2]).value = ws.cell(i, col_cost).value
                for x in range(len(col_list3)):
                    new_ws.cell(new_max_row + 1, col_list3[x][2]).value = ws.cell(i, col_list3[x][1]).value
                # print(ws.cell(i, 53).value)
        else:
            for x in range(len(col_list1)):
                new_ws.cell(new_max_row + 1, col_list1[x][2]).value = ws.cell(i, col_list1[x][1]).value
            for fee in col_list2:
                if fee[0] == "Charge Amount":
                    if ws.cell(i, col_old_charge_amount).value is not None:
                        new_ws.cell(new_max_row + 1, fee[2]).value = ws.cell(i, col_cost).value
                if fee[0] == "Extra Charge Amount":
                    if ws.cell(i, col_old_extra_charge_amount).value is not None:
                        new_ws.cell(new_max_row + 1, fee[2]).value = ws.cell(i, col_cost).value
            for x in range(len(col_list3)):
                new_ws.cell(new_max_row + 1, col_list3[x][2]).value = ws.cell(i, col_list3[x][1]).value
            # print(ws.cell(i, 53).value)

    # 计算价格
    new_ws.cell(1, len(col_list) + 1).value = "按报价表首重"  # price1
    new_ws.cell(1, len(col_list) + 2).value = "按报价表续重"  # price2
    new_ws.cell(1, len(col_list) + 3).value = "按报价表运费"  # price3
    new_ws.cell(1, len(col_list) + 4).value = "按报价表总金额"  # price4
    new_ws.cell(1, len(col_list) + 5).value = "旺季附加费"
    new_ws.cell(1, len(col_list) + 6).value = "人工重新打印运单（面单无法辨识，人工修正地址等）附加费"
    new_ws.cell(1, len(col_list) + 7).value = "超规，非规则/异形包裹的货物（Sperrgut）：a.长度大于120CM小于200CM b.非长方体 c.次长边或最短边大于60CM " \
                                              "d.重量大于31.5KG，会被拒收/退运并收费 "
    for i in range(2, new_ws.max_row + 1):
        if new_ws.cell(i, 13).value is None:
            continue
        if new_ws.cell(i, 13).value == '':
            continue
        weight = float(new_ws.cell(i, 13).value)
        price1 = 0
        price2 = 0
        price3 = 0
        price4 = 0

        # 客户代码
        customer_code = str(new_ws.cell(i, 5).value)
        if str(new_ws.cell(i, 5).value) == 'rong02-DHL':
            customer_code = "RONG"
        elif str(new_ws.cell(i, 5).value) == 'RGG-退货运单':
            customer_code = "RGG"

        # 根据发货渠道计算费用
        if str(new_ws.cell(i, 6).value).find("DHL德国非FBA") > -1:
            business_type = "DHL德国非FBA"
            price3 = query_data(data_price, business_type, customer_code, weight)[0]
            price4 = price3
        elif str(new_ws.cell(i, 6).value).find("DHL德国FBA") > -1:
            business_type = "DHL德国FBA"
            price3 = query_data(data_price, business_type, customer_code, weight)[0]
            price4 = price3
        elif str(new_ws.cell(i, 6).value).find("DHL欧盟主要国家") > -1:
            business_type = "DHL欧盟主要国家"
            weight_country = str(new_ws.cell(i, 7).value)

            price = query_data(data_price, business_type, customer_code, weight_country)
            price1 = float(price[0])
            price2 = float(price[1])
            price2 = (weight - 1) * price2
            price4 = price1 + price2

        new_ws.cell(i, len(col_list) + 1).value = price1
        new_ws.cell(i, len(col_list) + 2).value = price2
        new_ws.cell(i, len(col_list) + 3).value = price3
        new_ws.cell(i, len(col_list) + 4).value = price4

        progressbar['value'] = (i / new_ws.max_row) * 100
        root.update()
    try:
        wb.save(fullpath)
        wb.close()
    except PermissionError:
        tipps_value.set('文件已被占用,请先关闭')
        root.update()
        return 'please close file'

    return 'finish'
